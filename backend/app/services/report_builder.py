"""Build Excel reports for scenario export."""

from __future__ import annotations

import io
import logging
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.pipeline import Pipeline
from app.models.facility import Facility

logger = logging.getLogger(__name__)

# Gassco-style colors
NAVY_FILL = PatternFill(start_color="000D4D", end_color="000D4D", fill_type="solid")
TEAL_FILL = PatternFill(start_color="00D4AA", end_color="00D4AA", fill_type="solid")
HEADER_FILL = PatternFill(start_color="0A1628", end_color="0A1628", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
TEAL_FONT = Font(color="00D4AA", bold=True, size=11)
HEADER_FONT = Font(color="E8EDF5", bold=True, size=10)
DATA_FONT = Font(color="E8EDF5", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="1A3A5C"),
    right=Side(style="thin", color="1A3A5C"),
    top=Side(style="thin", color="1A3A5C"),
    bottom=Side(style="thin", color="1A3A5C"),
)


def _style_header_row(ws: Any, row: int, num_cols: int) -> None:
    """Apply header styling to a row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _style_data_cell(cell: Any) -> None:
    """Apply data cell styling."""
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="center")


def _auto_width(ws: Any) -> None:
    """Auto-fit column widths."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def build_excel_report(
    scenario_data: dict[str, Any],
    db: Session,
) -> bytes:
    """Build a multi-sheet Excel workbook for a scenario.

    Returns the workbook as bytes.
    """
    wb = Workbook()
    config = scenario_data.get("config", {})
    result = scenario_data.get("result", {})

    # ── Sheet 1: Summary ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"  # type: ignore[union-attr]
    ws.sheet_properties.tabColor = "00D4AA"  # type: ignore[union-attr]

    summary_rows = [
        ("CarbonBlend Scenario Report", ""),
        ("", ""),
        ("Scenario Name", scenario_data.get("name", "")),
        ("Description", scenario_data.get("description", "")),
        ("Scenario ID", scenario_data.get("id", "")),
        ("Created", scenario_data.get("created_at", "")),
        ("Report Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
        ("", ""),
        ("Source Field NPDID", scenario_data.get("source_field_npdid", "")),
        ("CO2 Content (mol%)", scenario_data.get("co2_mol_pct", "")),
        ("Gas Flow Rate (MSm3/d)", scenario_data.get("gas_flow_rate_mscm_d", "")),
        ("Strategy", config.get("strategy", "")),
        ("Target CO2 (mol%)", config.get("target_co2_mol_pct", "")),
        ("Target Terminals", ", ".join(config.get("target_terminals", []))),
        ("Storage Site", config.get("storage_site", "")),
    ]

    # Best pathway cost if available
    existing = result.get("existing_pathways", [])
    if existing:
        best = existing[0]
        summary_rows.append(("", ""))
        summary_rows.append(("OPTIMIZATION RESULTS", ""))
        summary_rows.append(("Best Pathway Cost (MUSD/yr)", best.get("total_cost_musd_yr", "")))
        summary_rows.append(("Best Pathway", best.get("name", "")))
        summary_rows.append(("CO2 Removed (MTPA)", best.get("co2_removed_mtpa", "")))
        summary_rows.append(("Pathways Found", len(existing)))

    for i, (label, value) in enumerate(summary_rows, start=1):
        cell_a = ws.cell(row=i, column=1, value=label)  # type: ignore[union-attr]
        cell_b = ws.cell(row=i, column=2, value=value)  # type: ignore[union-attr]
        if i == 1:
            cell_a.font = Font(color="00D4AA", bold=True, size=14)
        elif label and label.isupper():
            cell_a.font = TEAL_FONT
        else:
            cell_a.font = Font(color="8899BB", size=10)
        cell_b.font = DATA_FONT
    _auto_width(ws)  # type: ignore[arg-type]

    # ── Sheet 2: Pathways ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Pathways")
    ws2.sheet_properties.tabColor = "B8FFE1"

    pathway_headers = ["Rank", "Name", "Terminal", "Total Cost (MUSD/yr)", "CO2 Removed (MTPA)", "CO2 Stored (MTPA)", "Steps"]
    for col, header in enumerate(pathway_headers, start=1):
        ws2.cell(row=1, column=col, value=header)
    _style_header_row(ws2, 1, len(pathway_headers))

    pathways = result.get("existing_pathways", []) + result.get("bridge_pathways", [])
    for i, pw in enumerate(pathways, start=2):
        steps_desc = " -> ".join(
            f"{s.get('type', '')}@{s.get('location', '')}" for s in pw.get("steps", [])
        )
        values = [
            pw.get("rank", i - 1),
            pw.get("name", ""),
            pw.get("terminal", ""),
            pw.get("total_cost_musd_yr", 0),
            pw.get("co2_removed_mtpa", 0),
            pw.get("co2_stored_mtpa", 0),
            steps_desc,
        ]
        for col, val in enumerate(values, start=1):
            cell = ws2.cell(row=i, column=col, value=val)
            _style_data_cell(cell)
    _auto_width(ws2)

    # ── Sheet 3: Infrastructure ───────────────────────────────────────
    ws3 = wb.create_sheet("Infrastructure")
    ws3.sheet_properties.tabColor = "4A6FA5"

    infra_headers = ["Type", "Name", "From", "To", "Diameter (in)", "Medium", "Grouping"]
    for col, header in enumerate(infra_headers, start=1):
        ws3.cell(row=1, column=col, value=header)
    _style_header_row(ws3, 1, len(infra_headers))

    # Gather infrastructure used in pathways
    used_locations: set[str] = set()
    for pw in pathways:
        for step in pw.get("steps", []):
            loc = step.get("location", "")
            if loc:
                used_locations.add(loc.upper())

    pipelines = db.query(Pipeline).all()
    row = 2
    for pipe in pipelines:
        name_upper = (pipe.name or "").upper()
        if any(loc in name_upper or name_upper in loc for loc in used_locations if loc):
            values = [
                "Pipeline",
                pipe.name,
                pipe.from_facility,
                pipe.to_facility,
                pipe.diameter_inches,
                pipe.medium,
                pipe.main_grouping,
            ]
            for col, val in enumerate(values, start=1):
                cell = ws3.cell(row=row, column=col, value=val)
                _style_data_cell(cell)
            row += 1

    # Add facilities
    facilities = db.query(Facility).all()
    for fac in facilities:
        fac_upper = (fac.name or "").upper()
        if any(loc in fac_upper or fac_upper in loc for loc in used_locations if loc):
            values = [
                "Facility",
                fac.name,
                fac.belongs_to_name or "",
                "",
                "",
                fac.kind or "",
                "",
            ]
            for col, val in enumerate(values, start=1):
                cell = ws3.cell(row=row, column=col, value=val)
                _style_data_cell(cell)
            row += 1
    _auto_width(ws3)

    # ── Sheet 4: CO2 Analysis ─────────────────────────────────────────
    ws4 = wb.create_sheet("CO2 Analysis")
    ws4.sheet_properties.tabColor = "FCC419"

    co2_headers = ["Pathway", "Step", "Type", "Location", "CO2 In (mol%)", "CO2 Out (mol%)", "Cost (MUSD/yr)"]
    for col, header in enumerate(co2_headers, start=1):
        ws4.cell(row=1, column=col, value=header)
    _style_header_row(ws4, 1, len(co2_headers))

    row = 2
    for pw in pathways:
        for step in pw.get("steps", []):
            values = [
                pw.get("name", ""),
                step.get("type", ""),
                step.get("type", ""),
                step.get("location", ""),
                step.get("co2_in", 0),
                step.get("co2_out", 0),
                step.get("cost_musd_yr", 0),
            ]
            for col, val in enumerate(values, start=1):
                cell = ws4.cell(row=row, column=col, value=val)
                _style_data_cell(cell)
            row += 1
    _auto_width(ws4)

    # ── Sheet 5: Tariff Breakdown ─────────────────────────────────────
    ws5 = wb.create_sheet("Tariff Breakdown")
    ws5.sheet_properties.tabColor = "FFA94D"

    tariff_headers = ["Pathway", "Segment", "K Element", "U Element", "I Element", "O Element", "Total (NOK/Sm3)"]
    for col, header in enumerate(tariff_headers, start=1):
        ws5.cell(row=1, column=col, value=header)
    _style_header_row(ws5, 1, len(tariff_headers))

    row = 2
    for pw in pathways:
        tb = pw.get("tariff_breakdown", {})
        segments = tb.get("segments", [])
        for seg in segments:
            values = [
                pw.get("name", ""),
                seg.get("name", ""),
                seg.get("k", 0),
                seg.get("u", 0),
                seg.get("i", 0),
                seg.get("o", 0),
                seg.get("total", 0),
            ]
            for col, val in enumerate(values, start=1):
                cell = ws5.cell(row=row, column=col, value=val)
                _style_data_cell(cell)
            row += 1

        # Totals row
        if segments:
            values = [
                pw.get("name", ""),
                "TOTAL",
                "",
                "",
                "",
                "",
                tb.get("total_nok_sm3", 0),
            ]
            for col, val in enumerate(values, start=1):
                cell = ws5.cell(row=row, column=col, value=val)
                _style_data_cell(cell)
                if col >= 2:
                    cell.font = TEAL_FONT
            row += 1
    _auto_width(ws5)

    # ── Sheet 6: Risk Register ────────────────────────────────────────
    ws6 = wb.create_sheet("Risk Register")
    ws6.sheet_properties.tabColor = "FF6B6B"

    risk_headers = ["Category", "Description", "Likelihood", "Impact", "Mitigation"]
    for col, header in enumerate(risk_headers, start=1):
        ws6.cell(row=1, column=col, value=header)
    _style_header_row(ws6, 1, len(risk_headers))

    # Auto-generate basic risks
    co2_pct = scenario_data.get("co2_mol_pct", 0) or 0
    risks: list[dict[str, str]] = [
        {
            "category": "Technical",
            "description": "Pipeline CO2 spec exceedance during transient conditions",
            "likelihood": "medium" if co2_pct > 2.5 else "low",
            "impact": "high",
            "mitigation": "Install inline CO2 monitoring with automatic shutoff valves",
        },
        {
            "category": "Commercial",
            "description": "Tariff increases due to Gassled regulatory changes",
            "likelihood": "low",
            "impact": "medium",
            "mitigation": "Negotiate long-term transport agreements with tariff caps",
        },
        {
            "category": "Operational",
            "description": "Processing plant downtime reducing CO2 removal capacity",
            "likelihood": "medium",
            "impact": "high",
            "mitigation": "Maintain backup blend route; negotiate priority processing slots",
        },
        {
            "category": "Market",
            "description": "Gas price volatility affecting project economics",
            "likelihood": "high",
            "impact": "medium",
            "mitigation": "Hedge gas price exposure; diversify terminal markets",
        },
        {
            "category": "Regulatory",
            "description": "Changes to CO2 tax regime affecting removal economics",
            "likelihood": "medium",
            "impact": "high",
            "mitigation": "Monitor regulatory developments; maintain flexibility in CO2 strategy",
        },
    ]

    if co2_pct > 5:
        risks.append({
            "category": "Technical",
            "description": f"High CO2 field ({co2_pct} mol%) may require dedicated removal infrastructure",
            "likelihood": "high",
            "impact": "high",
            "mitigation": "Evaluate membrane vs amine scrubbing; consider phased capacity build",
        })

    for i, risk in enumerate(risks, start=2):
        values = [
            risk["category"],
            risk["description"],
            risk["likelihood"],
            risk["impact"],
            risk["mitigation"],
        ]
        for col, val in enumerate(values, start=1):
            cell = ws6.cell(row=i, column=col, value=val)
            _style_data_cell(cell)
    _auto_width(ws6)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
